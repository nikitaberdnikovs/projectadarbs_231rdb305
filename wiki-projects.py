import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook, load_workbook 

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

wb = load_workbook('wiki.xlsx')
ws = wb.active
url = str(ws['A' + '2'].value)
driver.get(url)
time.sleep(2)
name = driver.find_element(By.CSS_SELECTOR, 'span.mw-page-title-main').text
# print(driver.page_source)
elements = driver.find_elements(By.XPATH, '//p | //ul')
i = 4
j = 4
for element in elements:
    # Find all "a" elements within the "p" element
    a_elements = element.find_elements(By.TAG_NAME, 'a')
    for lnk in a_elements:
         all_links = str(lnk.get_attribute('href'))
         if url[0:30] in all_links:
              if str(lnk.get_attribute('title')):
                     j += 1
                     if j > 17:
                             ws['A' + str(i)].value = lnk.get_attribute('title')
                             ws['C' + str(i)].value = lnk.get_attribute('href')
                             i += 1
driver.quit()
wb.save(str(name) +'.xlsx')
